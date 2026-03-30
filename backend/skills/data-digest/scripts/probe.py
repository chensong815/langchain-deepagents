#!/usr/bin/env python3
"""Lightweight probe for the data-digest skill."""

from __future__ import annotations

import csv
import codecs
import gzip
import json
import subprocess
import sys
from collections import deque
from pathlib import Path
from typing import Any


FULL_SCAN_BYTES = 64 * 1024 * 1024
LINE_ESTIMATE_SAMPLE_BYTES = 256 * 1024
ENCODING_SAMPLE_BYTES = 20_000
JSON_FULL_PARSE_LIMIT_BYTES = 5 * 1024 * 1024
GZIP_DECOMPRESSED_PROBE_LIMIT_BYTES = 8 * 1024 * 1024
GZIP_DECOMPRESSED_CHUNK_BYTES = 64 * 1024
CSV_PREVIEW_ROWS = 6
JSON_PREVIEW_RECORDS = 3
JSONL_SCHEMA_SAMPLE_RECORDS = 200
JSONL_SCHEMA_SAMPLE_BYTES = 512 * 1024
TEXT_HEAD_LINES = 30
TAIL_LINES = 5
PREVIEW_CHARS = 2_000
GZIP_STREAMABLE_EXTENSIONS = {
    ".csv",
    ".tsv",
    ".json",
    ".jsonl",
    ".ndjson",
    ".txt",
    ".log",
    ".md",
    ".text",
    ".yaml",
    ".yml",
    ".ini",
    ".conf",
    ".toml",
}

_STRATEGY = {
    "small": "Full read - fits in one pass",
    "medium": "Full read with summary stats",
    "large": "Chunked summary",
    "huge": "Query-driven or sampled analysis",
}

_SIZE_ORDER = {
    "small": 0,
    "medium": 1,
    "large": 2,
    "huge": 3,
}


def probe_file(filepath: str) -> dict[str, Any]:
    path = Path(filepath)
    if not path.exists():
        return {"error": f"File not found: {filepath}"}

    extension = _normalized_extension(path)
    size_bytes = path.stat().st_size
    dispatch_extension = _dispatch_extension(extension)

    report: dict[str, Any] = {
        "filepath": str(path),
        "filename": path.name,
        "extension": extension,
        "size_bytes": size_bytes,
        "size_human": _human_size(size_bytes),
        "file_type": None,
        "row_count": None,
        "row_count_estimated": False,
        "column_count": None,
        "columns": None,
        "columns_sampled": None,
        "sheets": None,
        "line_count": None,
        "line_count_estimated": False,
        "encoding": None,
        "head_preview": None,
        "tail_preview": None,
        "size_category": None,
        "strategy_recommendation": None,
        "sub_reader": None,
        "warnings": [],
    }

    try:
        if extension.endswith(".gz") and dispatch_extension != extension:
            report["warnings"].append("Gzip-compressed file: probe is reading the decompressed stream.")

        if dispatch_extension in {".csv", ".tsv"}:
            _probe_tabular(path, dispatch_extension, report)
        elif dispatch_extension in {".xlsx", ".xlsm"}:
            _probe_xlsx(path, report)
        elif dispatch_extension == ".xls":
            _probe_xls(path, report)
        elif dispatch_extension == ".ods":
            _probe_ods(path, report)
        elif dispatch_extension in {".json", ".jsonl", ".ndjson"}:
            _probe_json_family(path, dispatch_extension, report)
        elif dispatch_extension in {".txt", ".log", ".md", ".text", ".yaml", ".yml", ".ini", ".conf", ".toml"}:
            _probe_text(path, report)
        else:
            _probe_unknown(path, report)
    except ModuleNotFoundError as exc:
        report["error"] = f"Missing optional dependency: {exc.name}"
        report["warnings"].append("Probe could not use an optional parser. Fall back to bounded manual inspection.")
    except Exception as exc:  # noqa: BLE001
        report["error"] = f"{type(exc).__name__}: {exc}"

    if not report["size_category"]:
        report["size_category"] = _size_category_from_bytes(size_bytes)
    if not report["strategy_recommendation"]:
        report["strategy_recommendation"] = _STRATEGY[report["size_category"]]

    return report


def _normalized_extension(path: Path) -> str:
    suffixes = [suffix.lower() for suffix in path.suffixes]
    if not suffixes:
        return ""
    if len(suffixes) >= 2 and suffixes[-1] == ".gz":
        return f"{suffixes[-2]}{suffixes[-1]}"
    return suffixes[-1]


def _dispatch_extension(extension: str) -> str:
    if extension.endswith(".gz"):
        base_extension = extension[:-3]
        if base_extension in GZIP_STREAMABLE_EXTENSIONS:
            return base_extension
    return extension


def _human_size(size_bytes: int) -> str:
    value = float(size_bytes)
    for unit in ("B", "KB", "MB", "GB"):
        if value < 1024 or unit == "GB":
            return f"{value:.1f} {unit}"
        value /= 1024
    return f"{value:.1f} TB"


def _detect_encoding(path: Path) -> str:
    try:
        import chardet
    except ImportError:
        return "utf-8"

    with _open_binary_stream(path) as handle:
        sample = handle.read(ENCODING_SAMPLE_BYTES)
    detected = chardet.detect(sample)
    return detected.get("encoding") or "utf-8"


def _is_gzip_path(path: Path) -> bool:
    return path.suffix.lower() == ".gz"


def _open_binary_stream(path: Path):
    if _is_gzip_path(path):
        return gzip.open(path, "rb")
    return path.open("rb")


def _open_text_stream(path: Path, encoding: str, *, errors: str = "replace", newline: str | None = None):
    if _is_gzip_path(path):
        return gzip.open(path, "rt", encoding=encoding, errors=errors, newline=newline)
    return path.open("r", encoding=encoding, errors=errors, newline=newline)


def _read_text_scan(path: Path, encoding: str, *, head_limit: int, tail_limit: int) -> dict[str, Any]:
    head: list[str] = []
    tail: deque[str] = deque(maxlen=tail_limit)

    if _is_gzip_path(path):
        return _read_gzip_text_scan(path, encoding, head_limit=head_limit, tail_limit=tail_limit)

    if path.stat().st_size <= FULL_SCAN_BYTES:
        line_count = 0
        with _open_text_stream(path, encoding, errors="replace") as handle:
            for raw_line in handle:
                clean = raw_line.rstrip("\r\n")
                if len(head) < head_limit:
                    head.append(clean)
                tail.append(clean)
                line_count += 1
        return {
            "line_count": line_count,
            "line_count_estimated": False,
            "head_lines": head,
            "tail_lines": list(tail),
        }

    head = _read_head_lines(path, encoding, head_limit)
    estimated_count = _estimate_line_count(path, encoding)
    return {
        "line_count": estimated_count,
        "line_count_estimated": True,
        "head_lines": head,
        "tail_lines": [],
    }


def _split_complete_lines(buffer: str) -> tuple[list[str], str]:
    lines = buffer.splitlines(keepends=True)
    if lines and not lines[-1].endswith(("\n", "\r")):
        pending = lines.pop()
    else:
        pending = ""
    return [line.rstrip("\r\n") for line in lines], pending


def _estimate_total_gzip_text_bytes(
    *,
    decompressed_bytes: int,
    compressed_consumed: int,
    compressed_total_bytes: int,
) -> int | None:
    if compressed_consumed <= 0 or compressed_consumed >= compressed_total_bytes:
        return None
    return max(decompressed_bytes, int(decompressed_bytes * compressed_total_bytes / compressed_consumed))


def _read_gzip_text_scan(path: Path, encoding: str, *, head_limit: int, tail_limit: int) -> dict[str, Any]:
    head: list[str] = []
    tail: deque[str] = deque(maxlen=tail_limit)
    decoder = codecs.getincrementaldecoder(encoding)(errors="replace")
    pending = ""
    line_count_sample = 0
    decompressed_bytes = 0
    compressed_consumed = 0
    compressed_total_bytes = path.stat().st_size
    reached_eof = False

    with gzip.open(path, "rb") as handle:
        while decompressed_bytes < GZIP_DECOMPRESSED_PROBE_LIMIT_BYTES:
            budget = min(GZIP_DECOMPRESSED_CHUNK_BYTES, GZIP_DECOMPRESSED_PROBE_LIMIT_BYTES - decompressed_bytes)
            chunk = handle.read(budget)
            if not chunk:
                reached_eof = True
                break

            decompressed_bytes += len(chunk)
            if handle.fileobj is not None:
                compressed_consumed = max(compressed_consumed, int(handle.fileobj.tell()))

            complete_lines, pending = _split_complete_lines(pending + decoder.decode(chunk))
            for clean in complete_lines:
                if len(head) < head_limit:
                    head.append(clean)
                tail.append(clean)
                line_count_sample += 1

        if reached_eof:
            complete_lines, pending = _split_complete_lines(pending + decoder.decode(b"", final=True))
            for clean in complete_lines:
                if len(head) < head_limit:
                    head.append(clean)
                tail.append(clean)
                line_count_sample += 1
            if pending:
                clean = pending.rstrip("\r\n")
                if len(head) < head_limit:
                    head.append(clean)
                tail.append(clean)
                line_count_sample += 1

    effective_size_bytes = decompressed_bytes
    if not reached_eof:
        estimated_total_bytes = _estimate_total_gzip_text_bytes(
            decompressed_bytes=decompressed_bytes,
            compressed_consumed=compressed_consumed,
            compressed_total_bytes=compressed_total_bytes,
        )
        if estimated_total_bytes is not None:
            effective_size_bytes = max(effective_size_bytes, estimated_total_bytes)

    if not head and pending:
        head.append(pending[:PREVIEW_CHARS])

    if reached_eof:
        return {
            "line_count": line_count_sample,
            "line_count_estimated": False,
            "head_lines": head,
            "tail_lines": list(tail),
            "effective_size_bytes": effective_size_bytes,
            "effective_size_estimated": False,
            "truncated": False,
        }

    if line_count_sample <= 0:
        line_count_estimate: int | None = None
    elif effective_size_bytes > decompressed_bytes and decompressed_bytes > 0:
        line_count_estimate = max(line_count_sample, int(line_count_sample * effective_size_bytes / decompressed_bytes))
    else:
        line_count_estimate = line_count_sample

    return {
        "line_count": line_count_estimate,
        "line_count_estimated": True,
        "head_lines": head,
        "tail_lines": [],
        "effective_size_bytes": effective_size_bytes,
        "effective_size_estimated": True,
        "truncated": True,
    }


def _read_head_lines(path: Path, encoding: str, limit: int) -> list[str]:
    lines: list[str] = []
    with _open_text_stream(path, encoding, errors="replace") as handle:
        for _ in range(limit):
            line = handle.readline()
            if not line:
                break
            lines.append(line.rstrip("\r\n"))
    return lines


def _estimate_line_count(path: Path, encoding: str) -> int | None:
    with _open_binary_stream(path) as handle:
        raw = handle.read(LINE_ESTIMATE_SAMPLE_BYTES)
    if not raw:
        return 0

    sample = raw.decode(encoding, errors="replace")
    line_breaks = sample.count("\n")
    if line_breaks == 0:
        return None

    ratio = path.stat().st_size / max(len(raw), 1)
    return max(1, int(line_breaks * ratio))


def _size_category_from_count(kind: str, count: int | None, size_bytes: int) -> str:
    if count is None:
        return _size_category_from_bytes(size_bytes)
    if kind == "rows":
        if count < 500:
            return "small"
        if count < 5_000:
            return "medium"
        if count < 100_000:
            return "large"
        return "huge"
    if count < 200:
        return "small"
    if count < 2_000:
        return "medium"
    if count < 50_000:
        return "large"
    return "huge"


def _size_category_from_bytes(size_bytes: int) -> str:
    if size_bytes < 1 * 1024 * 1024:
        return "small"
    if size_bytes < 20 * 1024 * 1024:
        return "medium"
    if size_bytes < 200 * 1024 * 1024:
        return "large"
    return "huge"


def _max_size_category(*categories: str) -> str:
    return max(categories, key=lambda item: _SIZE_ORDER[item])


def _size_category_from_count_and_bytes(kind: str, count: int | None, size_bytes: int) -> str:
    category = _max_size_category(
        _size_category_from_count(kind, count, size_bytes),
        _size_category_from_bytes(size_bytes),
    )
    low_count_threshold = 10 if kind == "rows" else 20
    if size_bytes > JSON_FULL_PARSE_LIMIT_BYTES and (count is None or count <= low_count_threshold):
        category = _max_size_category(category, "large")
    return category


def _compact_preview(lines: list[str]) -> str | None:
    if not lines:
        return None
    return "\n".join(lines)[:PREVIEW_CHARS]


def _relax_csv_field_limit() -> None:
    limit = sys.maxsize
    while limit >= 131_072:
        try:
            csv.field_size_limit(limit)
            return
        except OverflowError:
            limit //= 10


def _read_tabular_preview_rows(path: Path, encoding: str, separator: str) -> tuple[list[list[str]], str | None]:
    _relax_csv_field_limit()
    preview_rows: list[list[str]] = []

    try:
        with _open_text_stream(path, encoding, errors="replace", newline="") as handle:
            reader = csv.reader(handle, delimiter=separator)
            for index, row in enumerate(reader):
                preview_rows.append([cell[:120] for cell in row])
                if index >= CSV_PREVIEW_ROWS - 1:
                    break
        return preview_rows, None
    except csv.Error as exc:
        fallback_lines = _read_head_lines(path, encoding, CSV_PREVIEW_ROWS)
        preview_rows = [[cell[:120] for cell in line.split(separator)] for line in fallback_lines]
        return preview_rows, f"CSV preview fell back to naive delimiter splitting because parsing raised: {exc}"


def _count_tabular_records(path: Path, encoding: str, separator: str) -> tuple[int | None, str | None]:
    _relax_csv_field_limit()
    record_count = 0

    try:
        with _open_text_stream(path, encoding, errors="replace", newline="") as handle:
            reader = csv.reader(handle, delimiter=separator)
            for _ in reader:
                record_count += 1
    except csv.Error as exc:
        return (
            None,
            "Exact CSV row counting fell back to physical line counting because parsing raised: "
            f"{exc}. Quoted multiline records may make the row count an overestimate.",
        )

    return max(record_count - 1, 0), None


def _probe_tabular(path: Path, extension: str, report: dict[str, Any]) -> None:
    encoding = _detect_encoding(path)
    separator = "," if extension == ".csv" else "\t"
    text_scan = _read_text_scan(path, encoding, head_limit=TEXT_HEAD_LINES, tail_limit=TAIL_LINES)
    effective_size_bytes = int(text_scan.get("effective_size_bytes") or path.stat().st_size)

    preview_rows, preview_warning = _read_tabular_preview_rows(path, encoding, separator)

    header = preview_rows[0] if preview_rows else []
    row_count = None
    row_count_estimated = text_scan["line_count_estimated"]
    row_count_warning = None
    if not text_scan["line_count_estimated"]:
        row_count, row_count_warning = _count_tabular_records(path, encoding, separator)
        row_count_estimated = row_count is None
    if row_count is None and text_scan["line_count"] is not None:
        row_count = max(int(text_scan["line_count"]) - 1, 0)

    report["file_type"] = "CSV" if extension == ".csv" else "TSV"
    report["encoding"] = encoding
    report["separator"] = separator
    report["line_count"] = text_scan["line_count"]
    report["line_count_estimated"] = text_scan["line_count_estimated"]
    report["row_count"] = row_count
    report["row_count_estimated"] = row_count_estimated
    report["column_count"] = len(header)
    report["columns"] = header or None
    report["columns_sampled"] = False if header else None
    report["head_preview"] = preview_rows[1:] or preview_rows or None
    report["tail_preview"] = _compact_preview(text_scan["tail_lines"])
    report["size_category"] = _size_category_from_count_and_bytes("rows", row_count, effective_size_bytes)
    report["strategy_recommendation"] = _STRATEGY[report["size_category"]]
    report["sub_reader"] = "readers/dataframe/SKILL.md"
    if text_scan["line_count_estimated"]:
        if _is_gzip_path(path) and text_scan.get("truncated"):
            report["warnings"].append(
                "Row and line counts are estimated from a bounded decompressed sample because the gzip file is large."
            )
        else:
            report["warnings"].append("Row and line counts are estimated from a byte sample because the file is large.")
    if preview_warning:
        report["warnings"].append(preview_warning)
    if row_count_warning:
        report["warnings"].append(row_count_warning)


def _probe_xlsx(path: Path, report: dict[str, Any]) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    sheet_names = list(wb.sheetnames)
    details: dict[str, Any] = {}
    total_rows = 0

    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        preview_rows: list[list[str]] = []
        for row_index, row in enumerate(ws.iter_rows(values_only=True)):
            preview_rows.append([_stringify_cell(value) for value in row])
            if row_index >= CSV_PREVIEW_ROWS - 1:
                break

        header = preview_rows[0] if preview_rows else []
        data_rows = max((ws.max_row or 0) - 1, 0)
        total_rows += data_rows
        details[sheet_name] = {
            "row_count": data_rows,
            "column_count": ws.max_column,
            "header": header,
            "head_preview": preview_rows[1:] or None,
            "sheet_state": ws.sheet_state,
        }

    wb.close()

    report["file_type"] = "Excel (xlsx)"
    report["sheets"] = sheet_names
    report["sheet_details"] = details
    report["row_count"] = total_rows
    report["size_category"] = _size_category_from_count("rows", total_rows, path.stat().st_size)
    report["strategy_recommendation"] = _STRATEGY[report["size_category"]]
    report["sub_reader"] = "readers/excel/SKILL.md"


def _probe_xls(path: Path, report: dict[str, Any]) -> None:
    import xlrd

    workbook = xlrd.open_workbook(path, on_demand=True)
    sheet_names = workbook.sheet_names()
    details: dict[str, Any] = {}
    total_rows = 0

    for sheet_name in sheet_names:
        sheet = workbook.sheet_by_name(sheet_name)
        header = [_stringify_cell(value) for value in sheet.row_values(0)] if sheet.nrows else []
        head_preview: list[list[str]] = []
        for row_index in range(1, min(sheet.nrows, CSV_PREVIEW_ROWS)):
            head_preview.append([_stringify_cell(value) for value in sheet.row_values(row_index)])

        data_rows = max(sheet.nrows - 1, 0)
        total_rows += data_rows
        details[sheet_name] = {
            "row_count": data_rows,
            "column_count": sheet.ncols,
            "header": header,
            "head_preview": head_preview or None,
        }

    workbook.release_resources()

    report["file_type"] = "Excel (xls legacy)"
    report["sheets"] = sheet_names
    report["sheet_details"] = details
    report["row_count"] = total_rows
    report["size_category"] = _size_category_from_count("rows", total_rows, path.stat().st_size)
    report["strategy_recommendation"] = _STRATEGY[report["size_category"]]
    report["sub_reader"] = "readers/excel/SKILL.md"


def _probe_ods(path: Path, report: dict[str, Any]) -> None:
    import pandas as pd

    excel_file = pd.ExcelFile(path, engine="odf")
    details: dict[str, Any] = {}
    for sheet_name in excel_file.sheet_names:
        sample = pd.read_excel(excel_file, sheet_name=sheet_name, nrows=5, engine="odf")
        details[sheet_name] = {
            "row_count": None,
            "column_count": len(sample.columns),
            "header": [str(column) for column in sample.columns],
            "head_preview": sample.head(5).astype(str).to_dict(orient="records"),
        }

    report["file_type"] = "ODS"
    report["sheets"] = excel_file.sheet_names
    report["sheet_details"] = details
    report["size_category"] = _size_category_from_bytes(path.stat().st_size)
    report["strategy_recommendation"] = _STRATEGY[report["size_category"]]
    report["sub_reader"] = "readers/excel/SKILL.md"
    report["warnings"].append("ODS row counts are not available from the lightweight probe; inspect a target sheet next.")


def _probe_json_family(path: Path, extension: str, report: dict[str, Any]) -> None:
    if extension in {".jsonl", ".ndjson"}:
        _probe_jsonl(path, report)
    else:
        _probe_json(path, report)


def _probe_json(path: Path, report: dict[str, Any]) -> None:
    encoding = _detect_encoding(path)
    size_bytes = path.stat().st_size
    line_count, estimated, effective_size_bytes = _json_stream_metrics(path, encoding)
    effective_size_bytes = effective_size_bytes or size_bytes
    report["file_type"] = "JSON"
    report["encoding"] = encoding
    report["line_count"] = line_count
    report["line_count_estimated"] = estimated

    if effective_size_bytes <= JSON_FULL_PARSE_LIMIT_BYTES:
        with _open_text_stream(path, encoding, errors="replace") as handle:
            data = json.load(handle)

        if isinstance(data, list):
            report["json_type"] = "array"
            report["row_count"] = len(data)
            report["row_count_estimated"] = False
            key_union: list[str] = []
            if any(isinstance(item, dict) for item in data):
                key_union = sorted({key for item in data if isinstance(item, dict) for key in item.keys()})
            report["columns"] = key_union or None
            report["column_count"] = len(key_union) if key_union else None
            report["columns_sampled"] = False if key_union else None
            report["head_preview"] = json.dumps(
                data[:JSON_PREVIEW_RECORDS], ensure_ascii=False, indent=2, default=str
            )[:PREVIEW_CHARS]
        elif isinstance(data, dict):
            report["json_type"] = "object"
            keys = list(data.keys())
            preview = {key: data[key] for key in keys[:10]}
            report["columns"] = keys
            report["column_count"] = len(keys)
            report["columns_sampled"] = False if keys else None
            report["head_preview"] = json.dumps(preview, ensure_ascii=False, indent=2, default=str)[:PREVIEW_CHARS]
        else:
            report["json_type"] = type(data).__name__
            report["head_preview"] = repr(data)[:PREVIEW_CHARS]
    else:
        preview = _read_prefix(path, encoding, PREVIEW_CHARS)
        report["json_type"] = _guess_json_top_level(preview)
        report["head_preview"] = preview
        report["warnings"].append("Large JSON file: probe used a bounded preview instead of parsing the whole payload.")

    if estimated and _is_gzip_path(path):
        report["warnings"].append("JSON sizing used a bounded decompressed sample because the gzip file is large.")

    if isinstance(report["row_count"], int):
        report["size_category"] = _size_category_from_count_and_bytes("rows", report["row_count"], effective_size_bytes)
    elif effective_size_bytes > JSON_FULL_PARSE_LIMIT_BYTES and report["line_count"] in {None, 0, 1}:
        report["size_category"] = "huge" if effective_size_bytes >= 200 * 1024 * 1024 else "large"
    else:
        report["size_category"] = _size_category_from_count_and_bytes("lines", report["line_count"], effective_size_bytes)
    report["strategy_recommendation"] = _STRATEGY[report["size_category"]]
    report["sub_reader"] = "readers/json/SKILL.md"


def _collect_jsonl_preview_and_columns(
    path: Path,
    encoding: str,
    *,
    full_scan: bool,
) -> tuple[list[Any], set[str], bool]:
    preview_records: list[Any] = []
    columns: set[str] = set()
    consumed_bytes = 0
    valid_records = 0
    sampled = False

    with _open_text_stream(path, encoding, errors="replace") as handle:
        for raw_line in handle:
            consumed_bytes += len(raw_line.encode(encoding, errors="replace"))
            line = raw_line.strip()
            if not line:
                continue

            try:
                item = json.loads(line)
            except json.JSONDecodeError:
                if len(preview_records) < JSON_PREVIEW_RECORDS:
                    preview_records.append({"_raw": line[:240]})
                continue

            valid_records += 1
            if len(preview_records) < JSON_PREVIEW_RECORDS:
                preview_records.append(item)
            if isinstance(item, dict):
                columns.update(item.keys())

            if not full_scan and (
                valid_records >= JSONL_SCHEMA_SAMPLE_RECORDS or consumed_bytes >= JSONL_SCHEMA_SAMPLE_BYTES
            ):
                sampled = True
                break

    return preview_records, columns, sampled


def _probe_jsonl(path: Path, report: dict[str, Any]) -> None:
    encoding = _detect_encoding(path)
    text_scan = _read_text_scan(path, encoding, head_limit=TEXT_HEAD_LINES, tail_limit=TAIL_LINES)
    effective_size_bytes = int(text_scan.get("effective_size_bytes") or path.stat().st_size)
    full_scan_for_columns = (
        not text_scan["line_count_estimated"] and effective_size_bytes <= JSON_FULL_PARSE_LIMIT_BYTES
    )
    preview_records, columns, columns_sampled = _collect_jsonl_preview_and_columns(
        path,
        encoding,
        full_scan=full_scan_for_columns,
    )

    report["file_type"] = "JSONL"
    report["encoding"] = encoding
    report["line_count"] = text_scan["line_count"]
    report["line_count_estimated"] = text_scan["line_count_estimated"]
    report["row_count"] = text_scan["line_count"]
    report["row_count_estimated"] = text_scan["line_count_estimated"]
    report["columns"] = sorted(columns) or None
    report["column_count"] = len(columns) if columns else None
    report["columns_sampled"] = columns_sampled if columns else None
    report["head_preview"] = json.dumps(preview_records, ensure_ascii=False, indent=2, default=str)[:PREVIEW_CHARS]
    report["tail_preview"] = _compact_preview(text_scan["tail_lines"])
    report["size_category"] = _size_category_from_count_and_bytes("rows", text_scan["line_count"], effective_size_bytes)
    report["strategy_recommendation"] = _STRATEGY[report["size_category"]]
    report["sub_reader"] = "readers/json/SKILL.md"
    if text_scan["line_count_estimated"]:
        if _is_gzip_path(path) and text_scan.get("truncated"):
            report["warnings"].append(
                "JSONL row count is estimated from a bounded decompressed sample because the gzip file is large."
            )
        else:
            report["warnings"].append("JSONL row count is estimated from a byte sample because the file is large.")
    if columns_sampled:
        report["warnings"].append("JSONL columns are inferred from a bounded sample of records and may be incomplete.")


def _probe_text(path: Path, report: dict[str, Any]) -> None:
    encoding = _detect_encoding(path)
    text_scan = _read_text_scan(path, encoding, head_limit=TEXT_HEAD_LINES, tail_limit=TAIL_LINES)
    effective_size_bytes = int(text_scan.get("effective_size_bytes") or path.stat().st_size)
    prefix = "\n".join(text_scan["head_lines"])

    report["file_type"] = "Text"
    report["encoding"] = encoding
    report["line_count"] = text_scan["line_count"]
    report["line_count_estimated"] = text_scan["line_count_estimated"]
    report["head_preview"] = prefix[:PREVIEW_CHARS]
    report["tail_preview"] = _compact_preview(text_scan["tail_lines"])
    report["text_pattern"] = _detect_text_pattern(prefix)
    report["size_category"] = _size_category_from_count_and_bytes("lines", text_scan["line_count"], effective_size_bytes)
    report["strategy_recommendation"] = _STRATEGY[report["size_category"]]
    report["sub_reader"] = "readers/txt/SKILL.md"

    with _open_binary_stream(path) as handle:
        raw_sample = handle.read(4_096)
    if b"\x00" in raw_sample:
        report["warnings"].append("File contains NUL bytes and may be binary despite the extension.")
    if text_scan["line_count_estimated"]:
        if _is_gzip_path(path) and text_scan.get("truncated"):
            report["warnings"].append("Line count is estimated from a bounded decompressed sample because the gzip file is large.")
        else:
            report["warnings"].append("Line count is estimated from a byte sample because the file is large.")


def _probe_unknown(path: Path, report: dict[str, Any]) -> None:
    report["file_type"] = "unknown"
    try:
        magic = subprocess.check_output(["file", str(path)], text=True).strip()
    except Exception:  # noqa: BLE001
        magic = None

    report["file_magic"] = magic
    report["size_category"] = _size_category_from_bytes(path.stat().st_size)
    report["strategy_recommendation"] = "Unknown type - inspect with `file` and do a bounded manual read."

    if not magic:
        return

    lowered = magic.lower()
    if "json" in lowered:
        report["sub_reader"] = "readers/json/SKILL.md"
    elif "csv" in lowered:
        report["sub_reader"] = "readers/dataframe/SKILL.md"
    elif "text" in lowered:
        report["sub_reader"] = "readers/txt/SKILL.md"


def _guess_json_top_level(preview: str) -> str:
    stripped = preview.lstrip()
    if stripped.startswith("{"):
        return "object"
    if stripped.startswith("["):
        return "array"
    return "unknown"


def _json_stream_metrics(path: Path, encoding: str) -> tuple[int | None, bool, int | None]:
    if _is_gzip_path(path):
        text_scan = _read_gzip_text_scan(path, encoding, head_limit=1, tail_limit=0)
        return (
            text_scan["line_count"],
            text_scan["line_count_estimated"],
            int(text_scan.get("effective_size_bytes") or 0) or None,
        )

    if path.stat().st_size <= FULL_SCAN_BYTES:
        line_count = 0
        effective_size_bytes = 0
        with _open_text_stream(path, encoding, errors="replace") as handle:
            for raw_line in handle:
                line_count += 1
                effective_size_bytes += len(raw_line.encode(encoding, errors="replace"))
        return line_count, False, effective_size_bytes
    return _estimate_line_count(path, encoding), True, None


def _read_prefix(path: Path, encoding: str, limit_chars: int) -> str:
    with _open_text_stream(path, encoding, errors="replace") as handle:
        return handle.read(limit_chars)


def _detect_text_pattern(sample: str) -> str:
    lowered = sample.lower()
    if _looks_like_log(sample):
        return "log"
    if any(token in lowered for token in ("[section", "=", "export ", "{", "server:", "database:")):
        return "config-like"
    return "prose"


def _looks_like_log(sample: str) -> bool:
    import re

    patterns = (
        r"\d{4}-\d{2}-\d{2}[t ]\d{2}:\d{2}:\d{2}",
        r"\b(info|warn|warning|error|debug|fatal|trace)\b",
        r"\[[a-z]+\]",
    )
    return sum(bool(re.search(pattern, sample, flags=re.IGNORECASE)) for pattern in patterns) >= 2


def _stringify_cell(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    return text[:120]


def main() -> None:
    if len(sys.argv) != 2:
        print("Usage: python3 probe.py <filepath>")
        sys.exit(1)
    print(json.dumps(probe_file(sys.argv[1]), indent=2, ensure_ascii=False, default=str))


if __name__ == "__main__":
    main()
